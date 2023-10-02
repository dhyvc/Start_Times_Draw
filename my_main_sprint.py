#!/usr/bin/python
# -*- coding: utf-8 -*-

import argparse
import io

import openpyxl
import datetime
import random
import csv
# from fpdf import FPDF
from zipfile import ZipFile
from io import StringIO, BytesIO


kids_categories = ['ילדים זינוק', 'ילדות זינוק', 'קצרצר']
shorty_categories = ['D12S', 'D14S', 'H12S', 'H14S']
youth_categories = ['D16S', 'D18S', 'H16S', 'H18S']
adult1_categories = ['H21S', 'D-OpenS', 'H-OpenS']
adult2_categories = ['D21S']
adult3_categories = ['D35S', 'D40S', 'D45S', 'D50S', 'H35S', 'H40S', 'H45S', 'H50S', 'H55S', 'H60S']
adult4_categories = ['D55S', 'D60S', 'D65S', 'D75S', 'H65S', 'H70S', 'H75S', 'H80S', 'H85S', 'H90S']



""" Forest categories
kids_categories = ['ילדים זינוק', 'ילדות זינוק', 'קצרצר']
shorty_categories = ['D12', 'D14B', 'H12', 'H14B']
short_categories = ['D14A', 'D16B', 'H14A', 'H16B']
gold_categories = ['D65B', 'D75', 'H75', 'H80', 'H85', 'H90']
short_plus_women_categories = ['D21C', 'D40', 'D45', 'D50', 'D55', 'D60', 'D65A']
short_plus_men_categories = ['H50B', 'H60B', 'H65', 'H70']
medium_youth_categories = ['D16A', 'D18B', 'H16A', 'H18B']
medium_A_categories = ['H50A', 'H55', 'H60A']
medium_B_categories = ['D18A', 'D21B', 'D35', 'H21C', 'H35B', 'H45']
medium_plus_categories = ['D21A', 'H18A', 'H21B', 'H40']
long_categories = ['H21A', 'H35A']
"""

max_member_id = 10000
min_external_id = 20000
blank_slot_interval_minutes = 10
default_working_directory = './'


def sanity_check(first, last, period):
    if first is None:
        first = datetime.time(9, 0)
    if last is None:
        last = datetime.time(11, 0)
    if period is None:
        period = 30
    if first > last or period < 0:
        err = True
    else:
        err = False
    return first, last, period, err


def getperiods(competitors, first, last, period):
    periods = []
    for competitor in competitors:
        if competitor[5] < first:
            competitor[5] = first
        if competitor[5] > last:
            competitor[5] = last
    current_time = first
    minutes = current_time.minute
    hours = current_time.hour
    while current_time <= last:
        periods.append(current_time)
        if minutes + period > 59:
            hours += 1
        minutes = (minutes + period) % 60
        current_time = current_time.replace(hour=hours, minute=minutes)
    return periods, competitors


def read_start_file(filename, first, last, window_size, blank_slot_interval):
    """This function reads the startlist.xlsx file with the time allocation requests and returns data structures
    containing the relevant information for processing."""
    print('The input file name is: ' + filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    stno_col = sheet["A"]
    name_col = sheet["B"]
    club_col = sheet["C"]
    class_col = sheet["D"]
    requested_start_time_col = sheet["E"]
    start_time_col = sheet["F"]
    card_number_col = sheet["G"]
    phone_col = sheet["O"]
    competitors = []

    Shorty = []
    Youth = []
    Adults1 = []
    Adults2 = []
    Adults3 = []
    Kids = []
    Undefined = []
    courses = [Shorty, Youth, Adults1, Adults2, Adults3, Kids, Undefined]
    shorty_count = 0
    youth_count = 0
    adults1_count = 0
    adults2_count = 0
    adults3_count = 0
    kids_count = 0
    undefined_count = 0

""" Forest categories
    Shorty = []
    Short = []
    Gold = []
    Short_plus_women = []
    Short_plus_men = []
    Medium_youth = []
    Medium_A = []
    Medium_B = []
    Medium_plus = []
    Long = []
    Kids = []
    Undefined = []
    courses = [Shorty, Short, Gold, Short_plus_women, Short_plus_men, Medium_youth, Medium_A, Medium_B, Medium_plus,
               Long, Kids, Undefined]
    shorty_count = 0
    short_count = 0
    gold_count = 0
    short_plus_men_count = 0
    short_plus_women_count = 0
    medium_plus_count = 0
    medium_A_count = 0
    medium_B_count = 0
    medium_youth_count = 0
    long_count = 0
    kids_count = 0
    undefined_count = 0
"""

    for row in range(sheet.max_row):
        if row > 0:
            if class_col[row].value in shorty_categories:
                course = 'Shorty'
                shorty_count += 1
                Shorty.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                              value, requested_start_time_col[row].value, start_time_col[row].value,
                               card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in youth_categories:
                course = 'Youth'
                youth_count += 1
                Youth.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                             value, requested_start_time_col[row].value, start_time_col[row].value,
                              card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in gold_categories:
                course = 'Adults1'
                adults1_count += 1
                Adults1.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                            value, requested_start_time_col[row].value, start_time_col[row].value,
                             card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in short_plus_women_categories:
                course = 'Adults2'
                adults2_count += 1
                Adults2.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                         class_col[row].value, requested_start_time_col[row].value,
                                         start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in short_plus_men_categories:
                course = 'Adults3'
                adults3_count += 1
                Adults3.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                       class_col[row].value, requested_start_time_col[row].value,
                                       start_time_col[row].value, card_number_col[row].value, phone_col[row].value])

""" Forest categories
            if class_col[row].value in shorty_categories:
                course = 'Shorty'
                shorty_count += 1
                Shorty.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                              value, requested_start_time_col[row].value, start_time_col[row].value,
                               card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in short_categories:
                course = 'Short'
                short_count += 1
                Short.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                             value, requested_start_time_col[row].value, start_time_col[row].value,
                              card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in gold_categories:
                course = 'Gold'
                gold_count += 1
                Gold.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                            value, requested_start_time_col[row].value, start_time_col[row].value,
                             card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in short_plus_women_categories:
                course = 'Short_Plus_Women'
                short_plus_women_count += 1
                Short_plus_women.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                         class_col[row].value, requested_start_time_col[row].value,
                                         start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in short_plus_men_categories:
                course = 'Short_Plus_Men'
                short_plus_men_count += 1
                Short_plus_men.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                       class_col[row].value, requested_start_time_col[row].value,
                                       start_time_col[row].value, card_number_col[row].value, phone_col[row].value])

            elif class_col[row].value in medium_youth_categories:
                course = 'Medium_Youth'
                medium_youth_count += 1
                Medium_youth.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                     class_col[row].value, requested_start_time_col[row].value,
                                     start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in medium_A_categories:
                course = 'Medium_A'
                medium_A_count += 1
                Medium_A.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                 class_col[row].value, requested_start_time_col[row].value,
                                 start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in medium_B_categories:
                course = 'Medium_B'
                medium_B_count += 1
                Medium_B.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                 class_col[row].value, requested_start_time_col[row].value,
                                 start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in medium_plus_categories:
                course = 'Medium_Plus'
                medium_plus_count += 1
                Medium_plus.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                    class_col[row].value, requested_start_time_col[row].value,
                                    start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            elif class_col[row].value in long_categories:
                course = 'Long'
                long_count += 1
                Long.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                             class_col[row].value, requested_start_time_col[row].value,
                             start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
"""
            elif class_col[row].value in kids_categories:
                course = 'kids'
                kids_count += 1
                Kids.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                             class_col[row].value, requested_start_time_col[row].value,
                             start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            else:
                course = 'undefined'
                undefined_count += 1
                Undefined.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                  class_col[row].value, requested_start_time_col[row].value,
                                  start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
            competitors.append(
                [course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].value,
                 requested_start_time_col[row].value, start_time_col[row].value,
                 card_number_col[row].value, phone_col[row].value])
            competitors.sort(key=lambda x: x[0])
    for category in courses:
        blank_slot_counter = random.randint(1, 9)
        offset = 0
        category.sort(key=lambda x: x[5])
        periods, competitors = getperiods(competitors, first, last, window_size)
        runners_per_period = []
        for item in periods:
            runners_per_period.append([])
        for runner in category:
            for p in range(len(periods) - 1):
                if runner[5] <= periods[0]:
                    runners_per_period[0].append(runner)
                    break
                elif runner[5] >= periods[-1]:
                    runners_per_period[-1].append(runner)
                    break
                elif (runner[5] >= periods[p]) and (runner[5] < periods[p + 1]):
                    runners_per_period[p].append(runner)
                    break
            else:
                print("Error: Missing starting time.")
        next_vacant_slot = periods[0]
        ordered_starts = []
        for p in range(len(periods)):
            starts, next_vacant_slot, blank_slot_counter, offset = draw_start_times(p, periods, runners_per_period[p],
                                                                                    next_vacant_slot,
                                                                                    blank_slot_counter,
                                                                                    blank_slot_interval, offset)
            ordered_starts.append(starts)
        category.sort(key=lambda x: x[5])
    ordered_competitors = []
    for cat in courses:
        for runner in cat:
            ordered_competitors.append(runner)
    return ordered_competitors


def draw_start_times(current_window_index, start_windows, list_of_runners, first_open_slot, blank_slot_counter,
                     blank_slot_interval, offset):
    """This function accepts a list of runners, assigns each one a random number, sorts the runners according to
    the random number and assigns them a starting slot based on their order. Periodic vacancies will be inserted in
    order to support some flexibility for the organizers during the event."""

    # Determine the size of the interval required in order to balance the start time allocations around the requested time.
    number_of_runners = len(list_of_runners)
    balancing_offset = number_of_runners // 2
    # Establish the desired earliest time slot required for a balanced allocation.
    if start_windows[current_window_index].minute - balancing_offset < 0:
        balanced_first_start_hours = start_windows[current_window_index].hour - 1
        balanced_first_start_minutes = 60 - balancing_offset
    else:
        balanced_first_start_hours = start_windows[current_window_index].hour
        balanced_first_start_minutes = start_windows[current_window_index].minute - balancing_offset
    first_start_if_centered_around_requested_time = datetime.time(balanced_first_start_hours,
                                                                  balanced_first_start_minutes)
    # Pick the latest time between the next open slot and the earliest balanced time slot for this window. This will
    # serve as the next available time slot.
    next_open_slot = max(first_open_slot, first_start_if_centered_around_requested_time)
    # Assign each runner a random number.
    for runner in list_of_runners:
        runner.append(random.SystemRandom().random())
    list_of_runners.sort(key=lambda x: x[9])  # Sort the runners according to the random number assigned to them.
    if current_window_index == 0:  # First starting window - no balancing can be performed.
        next_open_slot = start_windows[0]
    elif current_window_index == len(start_windows) - 1:  # Last starting window - no balancing can be performed.
        earliest_needed_start_time = (datetime.datetime.combine(datetime.date.today(), start_windows[-1]) -
                                      datetime.timedelta(minutes=len(list_of_runners))).time()
        if earliest_needed_start_time < first_open_slot:
            next_open_slot = first_open_slot
        else:
            next_open_slot = earliest_needed_start_time
    offset = 0
    # blank_slot_counter = 1
    for runner in list_of_runners:
        if blank_slot_counter % blank_slot_interval == 0:
            offset += 1
            # next_open_slot = (datetime.datetime.combine(datetime.date.today(), next_open_slot) +
            #                   datetime.timedelta(minutes=1)).time()
        runner[5] = (datetime.datetime.combine(datetime.date.today(), next_open_slot) +
                     datetime.timedelta(minutes=list_of_runners.index(runner) + offset)).time()
        blank_slot_counter += 1
    next_open_slot = (datetime.datetime.combine(datetime.date.today(), next_open_slot) +
                      datetime.timedelta(minutes=len(list_of_runners) + offset)).time()
    return list_of_runners, next_open_slot, blank_slot_counter, offset


def write_start_file(competitors_list, working_dir):
    """This function writes out the startlist.csv file with starting times."""
    filename = working_dir + "StartList.csv"
    # print("Hello")
    # print("The output file path is:", filename)
    # file_in_memory = StringIO()
    # file_to_download = io.BytesIO()
    fields = ['STNO', 'NAME', 'CLUB', 'CLASS NAME', 'START TIME', 'CARD NUMBER', 'PHONE']
    startlist = []
    external_id = min_external_id
    competitors_list.sort(key=lambda x: x[1])
    for runner in competitors_list:
        if runner[1] > min_external_id:
            runner[1] = external_id
            external_id += 1
        startlist.append([runner[1], runner[2], runner[3], runner[4], runner[5], runner[7], runner[8]])
    rows = startlist
    # csv.writer(file_in_memory).writerow(fields)
    # csv.writer(file_in_memory).writerows(rows)
    # file_to_download.write(file_in_memory.getvalue().encode())
    # file_in_memory.close()
    # print(file_to_download)
    with open(filename, 'w', encoding='cp1255') as csvfile:
        # creating a csv writer object
        csvwriter = csv.writer(csvfile)

        # writing the fields
        csvwriter.writerow(fields)

        # writing the data rows
        csvwriter.writerows(rows)
    return filename


# def write_pdf_file(input_file):
#     pdf = FPDF('P', 'mm', 'A4')
#     pdf.add_page()
#     pdf.add_font('DejaVu', '', '/usr/share/fonts/TTF/DejaVuSansCondensed.ttf', uni=True)
#     pdf.set_font('DejaVu', '', 14)
#     # pdf.set_font('Arial', 'B', 16)
#     text = 'שלום לכם!'[::-1]
#     for x in range(40, 80, 10):
#         pdf.cell(40, x - 30, text)
#     pdf.output('tuto1.pdf', 'F')


def write_html_file_by_category(input_file, working_dir):
    starting_list_by_category = sorted(input_file, key=lambda x: (x[4], x[5]))
    start_list_file = open(working_dir + 'HTML_Start_Times_By_Category.html', 'w')
    start_list_file.write("""<html dir="rtl" lang="he">\n<head>\n<title> \nזמני זינוק \
           </title>\n</head> <body><h1><u>זמני זינוק</u></h1>\
           \n""")
    category = ''
    table_started = False
    for runner in starting_list_by_category:
        if runner[4] != category:
            if table_started:
                start_list_file.write("</table>")
            category = runner[4]
            table_started = True
            start_list_file.write("<H1>%s</H1>\n<table><tr><th>שעה</th><th></th><th>שם</th></tr>\n" % category)
        start_list_file.write("<tr><td>%s</td><td></td><td>%s</td></tr>\n" % (runner[5], runner[2]))
    start_list_file.write("</table>")
    start_list_file.write("</body>\n</html>")
    start_list_file.close()


def write_html_file_by_starting_time(input_file, working_dir):
    starting_list_by_category = sorted(input_file, key=lambda x: (x[5], x[4]))
    start_list_file = open(working_dir + 'HTML_Start_Times_By_Starting_Time.html', 'w')
    start_list_file.write("""<html dir="rtl" lang="he">\n<head>\n<title> \nרשימת זינוקים \
           </title>\n</head> <body><h1><u>רשימת זינוקים</u></h1>\
           \n""")
    time_slot = datetime.time()
    # time_slot = datetime.time(0, 0)
    table_started = False
    first_time_slot = True
    for runner in starting_list_by_category:
        if runner[5] != time_slot:
            if first_time_slot:
                first_time_slot = False
                previous_time_slot = runner[5]
            if table_started:
                start_list_file.write("</table>")
            # last_time_slot = time_slot
            category = runner[4]
            time_slot = runner[5]
            while (datetime.datetime.combine(datetime.date.today(), time_slot) -
                   datetime.datetime.combine(datetime.date.today(), previous_time_slot)) > datetime.timedelta(
                minutes=1):
                previous_time_slot = (datetime.datetime.combine(datetime.date.today(), previous_time_slot) +
                                      datetime.timedelta(minutes=1)).time()
                start_list_file.write("<H1>%s</H1>\n" % previous_time_slot)
            previous_time_slot = time_slot
            table_started = True
            start_list_file.write("<H1>%s</H1>\n<table border='1px'><tr><th></th><th>שם</th><th></th><th>קטגוריה</th"
                                  "><th></th><th>מספר SI</th><th></th><th>טלפון</th><th></th><th width='200px'>הערה</th></tr>\n" % time_slot)
        start_list_file.write("<tr><td><input type='checkbox'></td><td>%s</td><td></td><td>%s</td><td></td><td>%s</td"
                              "><td></td><td>%s</td><td></td><td></td></tr>\n" % (
                                  runner[2], runner[4], runner[7], runner[8]))
    start_list_file.write("</table>")
    start_list_file.write("</body>\n</html>")
    start_list_file.close()

def make_zip_file(directory, file_list):
    filename = directory + '/' + "StartList.zip"
    print(filename)
    # Create a ZipFile Object
    with ZipFile(filename, 'w') as zipObj2:
        # Add multiple files to the zip
        for f in file_list:
            zipObj2.write(directory + '/' + f)



# if __name__ == '__main__':
#     ap = argparse.ArgumentParser()
#     ap.add_argument("-i", "--inputfile", required=True, help="Input file name")
#     ap.add_argument("-fs", "--first_start", required=False, help="First start time")
#     ap.add_argument("-ls", "--last_start", required=False, help="Last start time")
#     ap.add_argument("-p", "--start_period", required=False, help="Start period size in minutes")
#     args = vars(ap.parse_args())
#     filepath = args['inputfile']
#     first_start = args['first_start']
#     last_start = args['last_start']
#     start_period_minutes = args['start_period']
#     first_start, last_start, start_period_minutes, error = sanity_check(first_start, last_start, start_period_minutes)
#     if error:
#         print("Error in the input parameters.")
#     event_competitors = read_start_file(filepath, first_start, last_start, start_period_minutes,
#                                         blank_slot_interval_minutes)
#     start_file = write_start_file(event_competitors, default_working_directory)
#     write_html_file_by_category(event_competitors, default_working_directory)
#     write_html_file_by_starting_time(event_competitors, default_working_directory)
    # write_pdf_file(event_competitors)
    pass
