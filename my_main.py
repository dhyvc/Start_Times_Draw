#!/usr/bin/python
# -*- coding: utf-8 -*-

import argparse
import io

import openpyxl
import datetime
import random
import csv
import os
# from fpdf import FPDF
from zipfile import ZipFile
from io import StringIO, BytesIO
from flask import render_template

kids_categories = ['ילדים זינוק', 'ילדות זינוק', 'קצרצר', 'ילדים כינוס', 'ילדות כינוס']
shorty_categories = ['D12', 'D14B', 'H12', 'H14B']
short_categories = ['D14A', 'D16B', 'H14A', 'H16B']
gold_categories = ['D65B', 'D75', 'D80', 'H75', 'H80', 'H85', 'H90']
short_plus_women_categories = ['D21C', 'D40', 'D45', 'D50', 'D55', 'D60', 'D65A']
short_plus_men_categories = ['H50B', 'H60B', 'H65', 'H70']
medium_youth_categories = ['D16A', 'D18B', 'H16A', 'H18B']
medium_A_categories = ['H50A', 'H55', 'H60A']
medium_B_categories = ['D18A', 'D21B', 'D35', 'H21C', 'H35B', 'H45']
medium_plus_categories = ['D21A', 'H18A', 'H21B', 'H40']
long_categories = ['H21A', 'H35A']
max_member_id = 15000
min_external_id = 20000
blank_slot_interval_minutes = 10
# consecutive_club_value = 1
# consecutive_category_value = 10
# optimization_iterations = 50
default_working_directory = './'


def sanity_check(first, last, period):
    if first is None:
        first = datetime.time(9, 0)
    if last is None:
        last = datetime.time(11, 0)
    if period is None:
        period = 30
    if first > last or period < 0:
        err = True
    else:
        err = False
    return first, last, period, err


def getperiods(competitors, first, last, period):
    """This function builds a list of starting slots based on the requested
    times, the window size, first start and last start. The function
    returns two lists: the starting slots and the competitors list."""
    periods = []
    # Sanity checks for requested starting times and starting period.
    for competitor in competitors:
        if competitor[5] < first:
            competitor[5] = first
        if competitor[5] > last:
            competitor[5] = last
    current_time = first
    minutes = current_time.minute
    hours = current_time.hour
    # Cycle through the possibles starting windows using the window size.
    while current_time <= last:
        # Add the current slot to the list.
        periods.append(current_time)
        # Update the next slot, advancing the hour as needed.
        if minutes + period > 59:
            hours += 1
        minutes = (minutes + period) % 60
        current_time = current_time.replace(hour=hours, minute=minutes)
    return periods, competitors

def quality_of_draw(runner_list, consecutive_category_value, consecutive_club_value):
    """This function calculates a score for a given draw based on consecutive
    runners from the same category and/or same club"""
    list_length = len(runner_list)
    print("BEGIN SCORING........")
    print("cat value:", consecutive_category_value)
    print("club value:", consecutive_club_value)
    if len(runner_list) > 0:
        print("category is:", runner_list[0][4])
        print("club is:", runner_list[0][3])
    print(runner_list)
    score = 0
    for i in range(list_length-1):
        if runner_list[i][3] == runner_list[i+1][3]:
            score += consecutive_club_value
            print("found consecutive club on indexes: ", i, i+1)
        if runner_list[i][4] == runner_list[i+1][4]:
            score += consecutive_category_value
            print("found consecutive categories on indexes: ", i, i+1)
    return score


def read_start_file(filename, first, last, window_size, blank_slot_interval, event_type, optimization_iterations, consecutive_club_value, consecutive_category_value):
    """This function reads the startlist.xlsx file with the time allocation requests and returns data structures
    containing the relevant information for processing."""
    print('The input file name is: ' + filename)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    stno_col = sheet["A"]
    name_col = sheet["B"]
    club_col = sheet["C"]
    class_col = sheet["D"]
    requested_start_time_col = sheet["E"]
    start_time_col = sheet["F"]
    card_number_col = sheet["G"]
    phone_col = sheet["P"]    
    # Determine the course to category mapping according to event type.
    
    if event_type == 'option2':
        kids_categories = ['ילדים זינוק', 'ילדות זינוק', 'קצרצר', 'ילדים כינוס', 'ילדות כינוס']
        shorty_categories = ['D12', 'D14B', 'H12', 'H14B']
        short_categories = ['קצר', 'D14A', 'D16B', 'H14A', 'H16B']
        gold_categories = ['D65B', 'D75', 'D80', 'H75', 'H80', 'H85', 'H90']
        short_plus_women_categories = ['D21C', 'D40', 'D45', 'D50', 'D55', 'D60', 'D65A']
        short_plus_men_categories = ['קצר+', 'קצר פלוס' , 'H50B', 'H60B', 'H65', 'H70']
        medium_youth_categories = ['בינוני', 'D16A', 'D18B', 'H16A', 'H18B']
        medium_A_categories = ['H50A', 'H55', 'H60A']
        medium_B_categories = ['D18A', 'D21B', 'D35', 'H21C', 'H35B', 'H45']
        medium_plus_categories = ['D21A', 'H18A', 'H21B', 'H40']
        long_categories = ['H21A', 'H35A']
        Shorty = []
        Short = []
        Gold = []
        Short_plus_women = []
        Short_plus_men = []
        Medium_youth = []
        Medium_A = []
        Medium_B = []
        Medium_plus = []
        Long = []
        Kids = []
        Undefined = []
        courses = [Shorty, Short, Gold, Short_plus_women, Short_plus_men, Medium_youth, Medium_A, Medium_B, Medium_plus, Long, Kids, Undefined]
        shorty_count = 0
        short_count = 0
        gold_count = 0
        short_plus_men_count = 0
        short_plus_women_count = 0
        medium_plus_count = 0
        medium_A_count = 0
        medium_B_count = 0
        medium_youth_count = 0
        long_count = 0
        kids_count = 0
        undefined_count = 0
    else:
        kids_categories = ['ילדים', 'ילדים זינוק', 'ילדות זינוק', 'קצרצר', 'ילדים לא תחר', 'ילדות לא תחר']
        shorty_categories = ['D12S', 'D14S', 'H12S', 'H14S']
        youth_categories = ['נוער', 'D16S', 'D18S', 'H16S', 'H18S']
        adult1_categories = ['H21S', 'D-OpenS', 'H-OpenS']
        adult2_categories = ['D21S', 'H35S', 'H40S', 'H45S']
        adult3_categories = ['D35S', 'D40S', 'D45S', 'D50S', 'H50S', 'H55S']
        adult4_categories = ['D55S', 'D60S', 'D65S', 'D75S', 'H60S', 'H65S', 'H70S', 'H75S', 'H80S', 'H85S', 'H90S']
        Shorty = []
        Youth = []
        Adults1 = []
        Adults2 = []
        Adults3 = []
        Adults4 = []
        Kids = []
        Undefined = []
        courses = [Shorty, Youth, Adults1, Adults2, Adults3, Adults4, Kids, Undefined]
        shorty_count = 0
        youth_count = 0
        adults1_count = 0
        adults2_count = 0
        adults3_count = 0
        adults4_count = 0
        kids_count = 0
        undefined_count = 0

        

    competitors = []
    for row in range(sheet.max_row):
        if row > 0:
            # For forest events.
            if event_type == 'option2':
                # Add the curren row data to the relevant course list.
                if class_col[row].value in shorty_categories:
                    course = 'Shorty'
                    shorty_count += 1
                    Shorty.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                                value, requested_start_time_col[row].value, start_time_col[row].value,
                                card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in short_categories:
                    course = 'Short'
                    short_count += 1
                    Short.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                                value, requested_start_time_col[row].value, start_time_col[row].value,
                                card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in gold_categories:
                    course = 'Gold'
                    gold_count += 1
                    Gold.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                                value, requested_start_time_col[row].value, start_time_col[row].value,
                                card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in short_plus_women_categories:
                    course = 'Short_Plus_Women'
                    short_plus_women_count += 1
                    Short_plus_women.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                            class_col[row].value, requested_start_time_col[row].value,
                                            start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in short_plus_men_categories:
                    course = 'Short_Plus_Men'
                    short_plus_men_count += 1
                    Short_plus_men.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                        class_col[row].value, requested_start_time_col[row].value,
                                        start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in medium_youth_categories:
                    course = 'Medium_Youth'
                    medium_youth_count += 1
                    Medium_youth.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                        class_col[row].value, requested_start_time_col[row].value,
                                        start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in medium_A_categories:
                    course = 'Medium_A'
                    medium_A_count += 1
                    Medium_A.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                    class_col[row].value, requested_start_time_col[row].value,
                                    start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in medium_B_categories:
                    course = 'Medium_B'
                    medium_B_count += 1
                    Medium_B.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                    class_col[row].value, requested_start_time_col[row].value,
                                    start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in medium_plus_categories:
                    course = 'Medium_Plus'
                    medium_plus_count += 1
                    Medium_plus.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                        class_col[row].value, requested_start_time_col[row].value,
                                        start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in long_categories:
                    course = 'Long'
                    long_count += 1
                    Long.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                class_col[row].value, requested_start_time_col[row].value,
                                start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in kids_categories:
                    course = 'kids'
                    kids_count += 1
                    Kids.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                class_col[row].value, requested_start_time_col[row].value,
                                start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                else:
                    # Check for missing class for competitor and add such competitor to the undefined list.
                    if not class_col[row].value:
                        class_col[row].value = "Missing class"
                    course = 'undefined'
                    undefined_count += 1
                    Undefined.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                    class_col[row].value, requested_start_time_col[row].value,
                                    start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                # Add the competitor to the competitors list.
                competitors.append(
                    [course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].value,
                    requested_start_time_col[row].value, start_time_col[row].value,
                    card_number_col[row].value, phone_col[row].value])
                # Sort the competitors list according to the course.
                competitors.sort(key=lambda x: x[0])
            else:
                # For sprint events.
                if class_col[row].value in shorty_categories:
                    # Add the row data to the relevant course list.
                    course = 'Shorty'
                    shorty_count += 1
                    Shorty.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                                value, requested_start_time_col[row].value, start_time_col[row].value,
                                card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in youth_categories:
                    course = 'Youth'
                    youth_count += 1
                    Youth.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                                value, requested_start_time_col[row].value, start_time_col[row].value,
                                card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in adult1_categories:
                    course = 'Adults1'
                    adults1_count += 1
                    Adults1.append([course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].
                                value, requested_start_time_col[row].value, start_time_col[row].value,
                                card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in adult2_categories:
                    course = 'Adults2'
                    adults2_count += 1
                    Adults2.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                            class_col[row].value, requested_start_time_col[row].value,
                                            start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in adult3_categories:
                    course = 'Adults3'
                    adults3_count += 1
                    Adults3.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                        class_col[row].value, requested_start_time_col[row].value,
                                        start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in adult4_categories:
                    course = 'Adults4'
                    adults4_count += 1
                    Adults4.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                        class_col[row].value, requested_start_time_col[row].value,
                                        start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                elif class_col[row].value in kids_categories:
                    course = 'kids'
                    kids_count += 1
                    Kids.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                class_col[row].value, requested_start_time_col[row].value,
                                start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                else:
                    # Check for missing class for competitor and add such competitor to the undefined list.
                    if not class_col[row].value:
                        class_col[row].value = "Missing class"
                    course = 'undefined'
                    undefined_count += 1
                    Undefined.append([course, stno_col[row].value, name_col[row].value, club_col[row].value,
                                    class_col[row].value, requested_start_time_col[row].value,
                                    start_time_col[row].value, card_number_col[row].value, phone_col[row].value])
                # Add the competitor to the competitors list.
                competitors.append(
                    [course, stno_col[row].value, name_col[row].value, club_col[row].value, class_col[row].value,
                    requested_start_time_col[row].value, start_time_col[row].value,
                    card_number_col[row].value, phone_col[row].value])
                # Sort the competitors list according to the course.
                competitors.sort(key=lambda x: x[0])

    for category in courses:
        # Establish the first blank slot for each course.
        blank_slot_counter = random.randint(1, 9)
        # Support bigger spacing between starts in the longer courses of a forest event.
        offset = 0
        # If the current course is long, then mark it as both long and medium so there will be two minutes added between starts.
        if (event_type == 'option2') and (category == Long):
            print ("long category found")
            long_category, medium_category = True, False
        # If the course is either medium youth, short, or medium plus, mark it as meduim category only for only 1 min added between starts.
        elif  (event_type == 'option2') and ((category == Medium_youth) or (category == Medium_plus) or (category == Short) or (category == Shorty)):
            print ("medium category found")
            long_category, medium_category = False, True
        else:
            long_category, medium_category = False, False
        # Sort the course according to the requested starting time parameter.
        category.sort(key=lambda x: x[5])
        # Retrieve both the starting slots list and the competitors list from the file downloaded from the ISOA site.
        periods, competitors = getperiods(competitors, first, last, window_size)
        # Assign runners to each time slot for the current course according to their requested starting time.
        runners_per_period = []
        for item in periods:
            runners_per_period.append([])
        for runner in category:
            for p in range(len(periods) - 1):
                if runner[5] <= periods[0]:
                    runners_per_period[0].append(runner)
                    break
                elif runner[5] >= periods[-1]:
                    runners_per_period[-1].append(runner)
                    break
                elif (runner[5] >= periods[p]) and (runner[5] < periods[p + 1]):
                    runners_per_period[p].append(runner)
                    break
            else:
                print("Error: Missing starting time.")
        # Establish the next possible starting time for the current course.
        next_vacant_slot = periods[0]
        ordered_starts = []
        # Cycle through the starting slots and randomly draw the starting order for each starting slot.
        for p in range(len(periods)):
#            print ('course name: ' + category[0][0])
            print ('course size: ' + str(len(category)))
            starts, next_vacant_slot, blank_slot_counter, offset = draw_start_times(p, periods, runners_per_period[p],
                                                                                    next_vacant_slot,
                                                                                    blank_slot_counter,
                                                                                    blank_slot_interval, offset, long_category, medium_category, optimization_iterations, consecutive_club_value, consecutive_category_value)
            ordered_starts.append(starts)
        category.sort(key=lambda x: x[5])
    ordered_competitors = []
    for cat in courses:
        for runner in cat:
            ordered_competitors.append(runner)
    return ordered_competitors


def draw_start_times(current_window_index, start_windows, list_of_runners, first_open_slot, blank_slot_counter,
                     blank_slot_interval, offset, long_category, medium_category, optimization_iterations, consecutive_club_value, consecutive_category_value):
    """This function accepts a list of runners, assigns each one a random number, sorts the runners according to
    the random number and assigns them a starting slot based on their order. Periodic vacancies will be inserted in
    order to support some flexibility for the organizers during the event. The long category gets special spacing of two minutes in forest events."""

    # Determine the size of the interval required in order to balance the start time allocations around the requested time.
    number_of_runners = len(list_of_runners)
    best_list_of_runners = []
    if long_category:
        additional_space = 2
    elif medium_category:
        additional_space = 1
    else:
        additional_space = 0
    balancing_offset = (1 + additional_space) * number_of_runners // 2
    if long_category:
        print ("balancing offset = " + str(balancing_offset))
    # Establish the desired earliest time slot required for a balanced allocation.
    if start_windows[current_window_index].minute - balancing_offset < 0:
        balanced_first_start_hours = start_windows[current_window_index].hour - 1
        balanced_first_start_minutes = max(60 - balancing_offset, 0)
    else:
        balanced_first_start_hours = start_windows[current_window_index].hour
        balanced_first_start_minutes = start_windows[current_window_index].minute - balancing_offset
    print ("balanced first start hours: " + str(balanced_first_start_hours))
    print ("balanced first start minutes: " + str(balanced_first_start_minutes))
    first_start_if_centered_around_requested_time = datetime.time(balanced_first_start_hours,
                                                                  balanced_first_start_minutes)
    print("first start if balanced: " + str(first_start_if_centered_around_requested_time))
    # Pick the latest time between the next open slot and the earliest balanced time slot for this window. This will
    # serve as the next available time slot.
    if number_of_runners > 0:
        next_open_slot = max(first_open_slot, first_start_if_centered_around_requested_time)
    else:
        next_open_slot = first_open_slot
    # Copy the list of runners into the "best" list
    for index in range (len(list_of_runners)):
        best_list_of_runners.append(list_of_runners[index])  # Establish base "best" draw.
     # Assign each runner a random number.
    for runner in best_list_of_runners:
        runner.append(random.SystemRandom().random())
    best_list_of_runners.sort(key=lambda x: x[9])  # Sort the runners according to the random number assigned to them.
    base_score = quality_of_draw (best_list_of_runners, consecutive_category_value, consecutive_club_value)  # Establish base quality score
    print("best draw list contains ", len(best_list_of_runners), "entries")
    print("and here they are:", best_list_of_runners)
    print("base score: ", base_score)
    # Repeat draw and see if we get a better draw with less consecutive runners from the same club and/or category
    for iter in range(optimization_iterations):
        for runner in list_of_runners:
            runner[9] = random.SystemRandom().random()
        list_of_runners.sort(key=lambda x: x[9])
        current_score = quality_of_draw(list_of_runners, consecutive_category_value, consecutive_club_value)
        print("current score: ", current_score)
        if current_score < base_score:
            base_score = current_score
            print('new low score!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ', base_score)
            for i in range(len(list_of_runners)):
                best_list_of_runners[i] = list_of_runners[i]
    print("\nThe winning list:\n", best_list_of_runners,"\n")
    if current_window_index == 0:  # First starting window - no balancing can be performed.
        next_open_slot = start_windows[0]
    elif current_window_index == len(start_windows) - 1:  # Last starting window - no balancing can be performed.
        earliest_needed_start_time = (datetime.datetime.combine(datetime.date.today(), start_windows[-1]) -
                                      datetime.timedelta(minutes=len(best_list_of_runners))).time()
        if earliest_needed_start_time < first_open_slot:
            next_open_slot = first_open_slot
        else:
            next_open_slot = earliest_needed_start_time
    offset = 0
    # blank_slot_counter = 1
    for runner in best_list_of_runners:
        if blank_slot_counter % blank_slot_interval == 0:
            offset += 1
            # next_open_slot = (datetime.datetime.combine(datetime.date.today(), next_open_slot) +
            #                   datetime.timedelta(minutes=1)).time()
        runner[5] = (datetime.datetime.combine(datetime.date.today(), next_open_slot) +
                     datetime.timedelta(minutes=best_list_of_runners.index(runner) * (1 + additional_space) + offset)).time()
        blank_slot_counter += 1
        if long_category:
            print ("***************************")
            print ("next time: " + str (runner[5]))
    next_open_slot = (datetime.datetime.combine(datetime.date.today(), next_open_slot) +
                      datetime.timedelta(minutes=len(best_list_of_runners) * (1 + additional_space) + offset)).time()
    return best_list_of_runners, next_open_slot, blank_slot_counter, offset


def write_start_file(competitors_list, working_dir):
    """This function writes out the startlist.csv file with starting times."""
    filename = working_dir + "StartList.csv"
    # print("Hello")
    # print("The output file path is:", filename)
    # file_in_memory = StringIO()
    # file_to_download = io.BytesIO()
    fields = ['STNO', 'NAME', 'CLUB', 'CLASS NAME', 'START TIME', 'CARD NUMBER', 'PHONE', 'CARD MEMO']
    startlist = []
    external_id = min_external_id
    competitors_list.sort(key=lambda x: x[1])
    for runner in competitors_list:
        if runner[1] > min_external_id:
            runner[1] = external_id
            external_id += 1
    # Add a "rental" tag to runners who have no SI assigned to them during registration.
        if runner[7] is None or runner[7] < 100000:
            rental = "rental"
        else:
            rental = ""
        startlist.append([runner[1], runner[2], runner[3], runner[4], runner[5], runner[7], runner[8], rental])
    rows = startlist
    # csv.writer(file_in_memory).writerow(fields)
    # csv.writer(file_in_memory).writerows(rows)
    # file_to_download.write(file_in_memory.getvalue().encode())
    # file_in_memory.close()
    # print(file_to_download)
    if not os.path.exists(working_dir):
# if the folder directory is not present
# then create it.
        os.makedirs(working_dir )

    with open(filename, 'w', encoding='cp1255') as csvfile:
        # creating a csv writer object
        csvwriter = csv.writer(csvfile)

        # writing the fields
        csvwriter.writerow(fields)

        # writing the data rows
        csvwriter.writerows(rows)
    return filename


def write_undefined_registrations(competitors_list, working_dir):
    """This function writes out the startlist.csv file with starting times."""
    filename = working_dir + "Undefined_Registrations.csv"
    # print("Hello")
    # print("The output file path is:", filename)
    # file_in_memory = StringIO()
    # file_to_download = io.BytesIO()
    fields = ['COURSE', 'STNO', 'NAME', 'CLUB', 'CLASS NAME', 'START TIME', 'CARD NUMBER', 'PHONE']
    startlist = []
    external_id = min_external_id
    competitors_list.sort(key=lambda x: x[1])
    for runner in competitors_list:
        if runner[1] > min_external_id:
            runner[1] = external_id
            external_id += 1
        if runner[0] == "undefined":
            print ("found an undefined runner!")
            print ("Name:", runner[2])
            startlist.append([runner[0], runner[1], runner[2], runner[3], runner[4], runner[5], runner[7], runner[8]])
    rows = startlist
    # csv.writer(file_in_memory).writerow(fields)
    # csv.writer(file_in_memory).writerows(rows)
    # file_to_download.write(file_in_memory.getvalue().encode())
    # file_in_memory.close()
    # print(file_to_download)
    with open(filename, 'w', encoding='cp1255') as csvfile:
        # creating a csv writer object
        csvwriter = csv.writer(csvfile)

        # writing the fields
        csvwriter.writerow(fields)

        # writing the data rows
        csvwriter.writerows(rows)
    return filename


# def write_pdf_file(input_file):
#     pdf = FPDF('P', 'mm', 'A4')
#     pdf.add_page()
#     pdf.add_font('DejaVu', '', '/usr/share/fonts/TTF/DejaVuSansCondensed.ttf', uni=True)
#     pdf.set_font('DejaVu', '', 14)
#     # pdf.set_font('Arial', 'B', 16)
#     text = 'שלום לכם!'[::-1]
#     for x in range(40, 80, 10):
#         pdf.cell(40, x - 30, text)
#     pdf.output('tuto1.pdf', 'F')


def write_html_file_for_website(input_file, working_dir):
    starting_list_by_category = sorted(input_file, key=lambda x: (x[4], x[5]))
    start_list_file = open(working_dir + 'startlist_web.html','w',encoding="utf-8")
    start_list_file.write(render_template('startlist.html', competitors=starting_list_by_category))
    start_list_file.close()

def write_html_file_by_category(input_file, working_dir):
    external_id = min_external_id
    input_file.sort(key=lambda x: x[1])
    for runner in input_file:
#        print ("old runner id: ", runner[1])
        if runner[1] > max_member_id:
            print ("old runner id: ", runner[1])
            runner[1] = external_id
            print ("new runner id: ", runner[1])
            external_id += 1
    
#    starting_list_by_category = sorted(input_file, key=lambda x: x[5])
#    starting_list_by_category = sorted(input_file, key=lambda x: x[4])
    starting_list_by_category = sorted(input_file, key=lambda x: (x[4], x[5]))
    start_list_file = open(working_dir + 'HTML_Start_Times_By_Category.html', 'w', encoding='utf8')
    

    
    if not os.path.exists(working_dir):
    # if the folder directory is not present
    # then create it.
        os.makedirs(working_dir )

    start_list_file.write("""<html dir="rtl" lang="he">\n<head>\n<meta charset="utf-8">\n<title> \nזמני זינוק \
           </title>\n</head> <body><h1><u>זמני זינוק</u></h1>\
           \n""")
    category = ''
    table_started = False
    for runner in starting_list_by_category:
        if runner[4] != category:
            if table_started:
                start_list_file.write("</table>")
            category = runner[4]
            table_started = True
            start_list_file.write("<H1>%s</H1>\n<table><tr><th align='center'>שעה</th><th align='center'>מספר חבר</th><th align='center'>שם</th><th></th><th align='center'>מועדון</th><th></th><th align='center'>SI</th></tr>\n" % category)
        start_list_file.write("<tr><td>%s</td><td align='center'>%s</td><td style='color: blue'><b>%s</b></td><td></td><td>%s</td><td></td><td align='center' style='color: blue'>%s</td></tr>\n" % (runner[5], runner[1], runner[2], runner[3], runner[7]))
    start_list_file.write("</table>")
    start_list_file.write("</body>\n</html>")
    start_list_file.close()


def write_html_file_by_starting_time(input_file, working_dir):
    starting_list_by_category = sorted(input_file, key=lambda x: (x[5], x[4]))
    start_list_file = open(working_dir + 'HTML_Start_Times_By_Starting_Time.html', 'w', encoding='utf8')
    start_list_file.write("""<html dir="rtl" lang="he">\n<head>\n<meta charset="utf-8">\n<title> \nרשימת זינוקים \
           </title>\n</head> <body><h1><u>רשימת זינוקים</u></h1>\
           \n""")
    time_slot = datetime.time()
    # time_slot = datetime.time(0, 0)
    table_started = False
    first_time_slot = True
    for runner in starting_list_by_category:
        if runner[5] != time_slot:
            if first_time_slot:
                first_time_slot = False
                previous_time_slot = runner[5]
            if table_started:
                start_list_file.write("</table>")
            # last_time_slot = time_slot
            category = runner[4]
            time_slot = runner[5]
            while (datetime.datetime.combine(datetime.date.today(), time_slot) -
                   datetime.datetime.combine(datetime.date.today(), previous_time_slot)) > datetime.timedelta(
                minutes=1):
                previous_time_slot = (datetime.datetime.combine(datetime.date.today(), previous_time_slot) +
                                      datetime.timedelta(minutes=1)).time()
                start_list_file.write("<H1>%s</H1>\n" % previous_time_slot)
            previous_time_slot = time_slot
            table_started = True
            start_list_file.write("<H1>%s</H1>\n<table border='1px'><tr><th></th><th>שם</th><th></th><th>קטגוריה</th"
                                  "><th></th><th>מספר SI</th><th></th><th>טלפון</th><th></th><th width='200px'>הערה</th></tr>\n" % time_slot)
        start_list_file.write("<tr><td><input type='checkbox'></td><td>%s</td><td></td><td>%s</td><td></td><td>%s</td"
                              "><td></td><td>%s</td><td></td><td></td></tr>\n" % (
                                  runner[2], runner[4], runner[7], runner[8]))
    start_list_file.write("</table>")
    start_list_file.write("</body>\n</html>")
    start_list_file.close()
    
def write_vacant_slots_by_course(input_file, working_dir, first_start, last_start, event_type):
    """This function writes out the Vacancies.xlsx file with vacant starting times."""
    filename = working_dir + "Vacancies.csv"
    xlfilename = working_dir + "Vacancies.xlsx"
    
    # Establish the category list row height according to the event type.
    forrest_categories_row_height = 50
    sprint_categories_row_height = 80
    
    if event_type == 'option2':
        course_fields = ['Shorty', 'Short', 'Gold', 'Short_Plus_Men', 'Short_Plus_Women', 'Medium_Youth', 'Medium_A', 'Medium_B', 'Medium_Plus', 'Long']
    
        category_fields = ['D12 D14B H12 H14B', 'D14A D16B H14A H16B קצר', 'D65B D75 D80 H75 H80 H85 H90', 'H50B H60B H65 H70 +קצר', 'D21C D40 D45 D50 D55 D60 D65A', 'D16A D18B H16A H18B בינוני', 'H50A H55 H60A', 'D18A D21B D35 H21C H35B H45', 'D21A H18A H21B H40', 'H21A H35A']
    else:
        course_fields = ['Shorty', 'Youth', 'Adults1', 'Adults2', 'Adults3', 'Adults4']
    
        category_fields = ['D12S D14S H12S H14S', 'D16S D18S H16S H18S נוער', 'H21S D-OpenS H-OpenS', 'D21S H35S H40S H45S', 'D35S D40S D45S D50S H50S H55S', 'D55S D60S D65S D75S H60S H65S H70S H75S H80S H85S H90S']
        
    male_color_palette = ['dd6727', 'd78644', '6088e1', 'd7ebf2', '679cde', '5c7ee2', '6aa6dc', '5260e7', '6392df', '5974e4', '78ced6', '75c4d8',  '556ae5', '6eb0db', '9fcddc', '09679b', '402eee', 'ffff00']
    femmale_color_palette = ['ff72a8', 'ff8fe9', 'ff8ade', 'ff44ff', 'ff81c9', 'ff00be', 'ff77b3', 'ff94f4',  'ff6e9e', 'ff6993', 'ff5567',  'ff5f7d', 'ff5a72', 'ff6488', 'ff515d', 'ff4c52', 'ff4747', 'ffff00']
    age_scale = [12, 14, 16, 18, 21, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90]
        
    
    # Housekeeping code for setting the page orientation, print settings
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vacancies"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.print_title_rows = '1:3'
    
    # Define border types
    thick_red = openpyxl.styles.Side(border_style="thick", color="FF0000")
    thick_black = openpyxl.styles.Side(border_style="thick", color="FFFFFF")
    thin_black = openpyxl.styles.Side(border_style="thin", color="000000")
    thick_red_cell_border_template = openpyxl.styles.Border(bottom = thick_red, top = thick_red, left = thick_red, right = thick_red)
    thin_black_cell_border_template = openpyxl.styles.Border(bottom = thin_black, top = thin_black, left = thin_black, right = thin_black)

    # Set up the title row
    title_cell = ws.cell(row =1, column = 1)
    title_cell.font = openpyxl.styles.Font(size = 14)
    title_cell.border = thick_red_cell_border_template
    title_cell.alignment = openpyxl.styles.Alignment(horizontal = 'center')
    ws.cell(row = 1, column = 1, value = "חלונות זינוק פנויים")
    
    # Set up the additional titles on the page
    ws.cell(row = 2, column = 1).border = thin_black_cell_border_template
    ws.cell(row = 2, column = 1).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
    ws.cell(row = 2, column = 1, value = "שעת זינוק")
    
    # Fill in the course row
    for index in range(len(course_fields)):
        ws.cell(row = 2, column=index + 2, value=course_fields[index])
        ws.cell(row = 2, column=index + 2).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
        ws.cell(row = 2, column=index + 2).border = thin_black_cell_border_template
    
    # Fill in the categories row
    for index in range(len(course_fields)):
        ws.cell(row = 3, column=index + 2, value=category_fields[index])
        ws.cell(row = 3, column=index + 2).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
        ws.cell(row = 3, column=index + 2).border = thin_black_cell_border_template
        
    # Fix the row height to accomodate the category list
    if event_type == 'option2':
        ws.row_dimensions[3].height = forrest_categories_row_height
    else:
        ws.row_dimensions[3].height = sprint_categories_row_height
        
    
    # Adjust the column width to include the labels with some spacing
    column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].bestFit = True
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].width *= 1.3
    ws.column_dimensions['A'].width /= 1.5
    
    # Sort the competitor list according to the start times
    
    input_file.sort(key = lambda x: x[5])
    print(input_file[0])
    print(input_file[10])
    print(input_file[20])
    print("last start time: ", input_file[-1][5])
    print("last start time type: ", type(input_file[-1][5]))
    last_start_time = input_file[-1][5]
    last_assigned_start_in_timedelta = datetime.datetime.combine(datetime.datetime.min, last_start_time) - datetime.datetime.min
    print("last assigned time in timedelta: ", last_assigned_start_in_timedelta)
    print("last start time in delta type: ", type(last_assigned_start_in_timedelta))
    print("first start: ", first_start)
    print("first start time type: ", type(first_start))
    first_assigned_start_in_timedelta = datetime.datetime.combine(datetime.datetime.min, first_start) - datetime.datetime.min
    print("first start in delta type: ", first_assigned_start_in_timedelta)
    print("first start time in delta type: ", type(first_assigned_start_in_timedelta))
    used_time_slots_in_minutes = (last_assigned_start_in_timedelta - first_assigned_start_in_timedelta).total_seconds() / 60
    print("used time slots: ", used_time_slots_in_minutes)
    
    # Prepare the starting times list in the format that allows filling the start times column
    delta = datetime.timedelta(minutes = 1)
    first_start_in_timedelta = datetime.timedelta(hours=first_start.hour, minutes=first_start.minute, seconds=first_start.second)
    last_start_in_timedelta = datetime.timedelta(hours=last_start.hour, minutes=last_start.minute, seconds=last_start.second)
    difference_delta = last_start_in_timedelta - first_start_in_timedelta
    current_time = first_start_in_timedelta
    print("minutes =: " + str(int(difference_delta.total_seconds() //60)))

    print (first_start_in_timedelta + delta)
    
    if used_time_slots_in_minutes < 120:
        number_of_rows = 120
    else:
        number_of_rows = used_time_slots_in_minutes
    
    # Fill the start times column
    for index in range (int(number_of_rows) +1):
#    for index in range (int(difference_delta.total_seconds() //60) +1):
        ws.cell(row = index + 4, column = 1, value = current_time)
        ws.cell(row = index + 4, column = 1).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center')
        ws.cell(row = index + 4, column = 1).border = thin_black_cell_border_template
        current_time += delta
        
    # Create borders for all the relevant cells in the worksheet
#    for line in range(4, len(input_file)+1):
    for line in range(4, int(number_of_rows) +5):
        for col in range(2, len(course_fields)+2):
            ws.cell(row = line, column = col).border = thin_black_cell_border_template
    
    # Loop over the starting times and the competitor lists and fill the cells that are occupied
    
    for index in range (int(number_of_rows) +1):
        for competitor in input_file:
            # Convert competitor start time to timedelta object
            competitor_start_time = datetime.datetime.combine(datetime.date.min, competitor[5]) - datetime.datetime.min
            if competitor[0] != "kids" and competitor[0] != "undefined":
                if competitor_start_time == ws.cell(row = index + 4, column = 1).value:
#                    print (competitor[0], competitor[5])
                    name = competitor[2]
                    comp_class = str(competitor[4])
                    try:
                        age = int(str(competitor[4])[1:3])
                        index_of_color = age_scale.index(age)
                    except:
                        index_of_color = 17
#                    print ('index of color: ' + str(index_of_color))
                    if str(competitor[4])[0] == 'D':
                        pallette = femmale_color_palette
                    else:
                        pallette = male_color_palette
                    ws.cell(row = index + 4, column = course_fields.index(competitor[0]) + 2, value = name + "[" + comp_class + "]")
                    ws.cell(row = index + 4, column = course_fields.index(competitor[0]) + 2).fill = openpyxl.styles.PatternFill(start_color=pallette[index_of_color], end_color=pallette[index_of_color], fill_type="solid")
    
    wb.save(xlfilename)



def make_zip_file(directory, file_list):
    filename = directory + '/' + "StartList.zip"
    print(filename)
    # Create a ZipFile Object
    with ZipFile(filename, 'w') as zipObj2:
        # Add multiple files to the zip
        for f in file_list:
            zipObj2.write(directory + '/' + f)

